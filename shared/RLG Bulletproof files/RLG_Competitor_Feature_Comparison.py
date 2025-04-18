#!/usr/bin/env python3
"""
RLG_Competitor_Feature_Comparison.xlsx Generator
-------------------------------------------------

This script generates a comprehensive competitor feature comparison Excel file for
RLG Data and RLG Fans. It is designed to be:
  - Robust, updated, and refined
  - Fully automated and data-driven
  - Scalable and user-friendly with region, country, city, and town accuracy
  - Informative regarding our scraping and compliance tools, all services, and the RLG Super Tool

The generated Excel workbook includes two sheets:
  1. Competitor Comparison:
     - Contains a detailed table comparing our tool against competitors such as Brandwatch,
       Brand24, Mention, Sprout Social, BuzzSumo, Keyhole, Hootsuite, Meltwater, Digimind,
       Brandmentions, Google Alerts, and our own RLG Data and RLG Fans.
  2. Tool Details & Recommendations:
     - Provides in-depth information about our scraping tools, compliance engine, the RLG Super Tool,
       our overall value proposition, and additional recommendations for future enhancements.

All sections are humanized and personalized to reflect the competitive value and strengths
of our platform.

Note: Replace any dummy data with live data or dynamically retrieved information as required.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_competitor_feature_comparison(filename="RLG_Competitor_Feature_Comparison.xlsx"):
    """
    Generates the Competitor Feature Comparison Excel file.

    Args:
        filename (str): The name of the output Excel file.
    """
    # Create a new workbook and rename the default sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competitor Comparison"
    
    # Define header row for the competitor comparison table
    headers = [
        "Competitor Name",
        "Price Competitiveness (1-10)",
        "Feature Set (1-10)",
        "Data Accuracy (1-10)",
        "Regional Coverage (1-10)",
        "User-Friendliness (1-10)",
        "Compliance & Security (1-10)",
        "Scraping Tools Integration (1-10)",
        "RLG Super Tool Value (1-10)",
        "Overall Score (1-10)",
        "Additional Comments"
    ]
    
    # Set header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B8BBE", end_color="4B8BBE", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dummy competitor data – replace with real data if available.
    competitor_data = [
        # Format: [Competitor, Price, Feature, Data Accuracy, Regional Coverage, User-Friendliness,
        # Compliance & Security, Scraping Integration, RLG Super Tool Value, Overall Score, Comments]
        ["Brandwatch",       8, 9, 9, 9, 8, 9, 8, 5, 8, "Excellent analysis; however, RLG Super Tool integration is limited."],
        ["Brand24",         7, 8, 8, 8, 7, 8, 7, 4, 7, "Solid tool with moderate regional coverage."],
        ["Mention",         7, 7, 8, 7, 8, 7, 7, 4, 7, "User-friendly but basic in advanced features."],
        ["Sprout Social",   6, 9, 8, 8, 9, 8, 7, 3, 8, "Great social media management; lower scraping tool focus."],
        ["BuzzSumo",        8, 8, 8, 7, 7, 7, 6, 3, 7, "Ideal for content; lacks in compliance automation."],
        ["Keyhole",         7, 7, 7, 8, 8, 7, 8, 3, 7, "Reliable but pricing may deter smaller enterprises."],
        ["Hootsuite",       6, 8, 8, 8, 9, 8, 7, 3, 8, "Comprehensive with decent scraping integration."],
        ["Meltwater",       8, 9, 9, 9, 8, 9, 8, 4, 8, "High-end solution with extensive media reach."],
        ["Digimind",        7, 8, 8, 8, 7, 8, 7, 4, 7, "Balanced approach to competitive intelligence."],
        ["Brandmentions",   7, 7, 7, 7, 7, 7, 7, 3, 7, "Adequate for basic monitoring tasks."],
        ["Google Alerts",  10, 5, 6, 6, 6, 6, 5, 2, 6, "Free tool; lacks depth and automation."],
        ["RLG Data and RLG Fans", 9, 10, 10, 10, 10, 10, 10, 10, 10, 
         "Our solution offers comprehensive, automated scraping, rigorous compliance, and seamless RLG Super Tool integration, delivering unmatched regional accuracy and data-driven insights."]
    ]
    
    # Write competitor data to the sheet
    for row_idx, row_data in enumerate(competitor_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths for better readability
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 20
    
    # ------------------------------------------------------------------------------
    # Create a second sheet for detailed tool information and recommendations
    ws_details = wb.create_sheet(title="Tool Details & Recommendations")
    
    # Title for the details sheet
    ws_details.merge_cells('A1:D1')
    ws_details['A1'] = "RLG Data and RLG Fans - Tool Details & Recommendations"
    ws_details['A1'].font = Font(bold=True, size=16)
    ws_details['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Define the details sections: Title and Content pairs
    details = [
        ("Scraping Tools", 
         "Our scraping tools automatically collect real-time data from diverse global sources. "
         "They ensure that our analysis reflects up-to-date regional, country, city, and town insights, "
         "providing a competitive edge through data precision."),
        
        ("Compliance Tools", 
         "Our compliance engine leverages state-of-the-art algorithms to continuously monitor regulatory standards "
         "(e.g., WCAG, GDPR, CCPA). This guarantees that all data-driven decisions are secure and meet the highest compliance benchmarks."),
        
        ("RLG Super Tool", 
         "The RLG Super Tool integrates data from our scraping and compliance modules into one unified, robust dashboard. "
         "It is scalable, user-friendly, and delivers actionable insights in real-time, ensuring your business remains ahead of the competition."),
        
        ("Value Proposition", 
         "By combining advanced scraping technology, rigorous compliance monitoring, and the innovative RLG Super Tool, "
         "our platform provides an unparalleled, automated solution. It is personalized and relatable, offering tailored insights "
         "that empower businesses to meet and exceed global accessibility standards."),
        
        ("Additional Recommendations", 
         "1. Continuously refine scraping algorithms to adapt to emerging data sources.\n"
         "2. Enhance UI personalization to address regional differences and user preferences.\n"
         "3. Integrate machine learning for predictive analysis and proactive issue detection.\n"
         "4. Expand regional granularity to include more localized data (cities, towns).\n"
         "5. Maintain strict security and compliance protocols to build long-term user trust.")
    ]
    
    # Write the details into the second sheet starting from row 3
    current_row = 3
    for title, content in details:
        ws_details.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=14)
        current_row += 1
        ws_details.cell(row=current_row, column=1, value=content).alignment = Alignment(wrapText=True)
        current_row += 2
    
    # Adjust column widths in the details sheet for clarity
    for col in range(1, 5):
        ws_details.column_dimensions[get_column_letter(col)].width = 40
    
    # ------------------------------------------------------------------------------
    # Save the workbook
    wb.save(filename)
    print(f"Competitor Feature Comparison Excel file generated: {filename}")

if __name__ == "__main__":
    generate_competitor_feature_comparison()
